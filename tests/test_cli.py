from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from gentms.cli import main


def test_shorthand_annual_command_creates_one_workbook(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)

    exit_code = main(["2026"])

    assert exit_code == 0
    output_path = tmp_path / "2026年度(R8).xlsx"
    assert output_path.exists()
    workbook = load_workbook(output_path)
    assert workbook.sheetnames[0] == "4月"


def test_monthly_command_creates_requested_files_in_given_order(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)

    exit_code = main(["monthly", "2026", "12", "1", "3"])

    assert exit_code == 0
    assert (tmp_path / "2026-12.xlsx").exists()
    assert (tmp_path / "2027-01.xlsx").exists()
    assert (tmp_path / "2027-03.xlsx").exists()


def test_create_dir_places_outputs_under_named_directory(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)

    exit_code = main(["monthly", "2026", "12", "--create-dir"])

    assert exit_code == 0
    target_dir = tmp_path / "2026年度(R8)"
    assert target_dir.exists()
    assert (target_dir / "2026-12.xlsx").exists()


def test_existing_output_returns_error(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    (tmp_path / "2026年度(R8).xlsx").write_text("occupied", encoding="utf-8")

    exit_code = main(["annual", "2026"])

    assert exit_code == 2


def test_duplicate_months_return_error(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)

    exit_code = main(["monthly", "2026", "12", "12"])

    assert exit_code == 2
