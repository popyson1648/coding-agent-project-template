from __future__ import annotations

from pathlib import Path
from typing import Iterable, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from gentms.fiscal import FISCAL_MONTHS, iter_month_dates, sheet_name

HEADERS = (
    "日付",
    "出勤 (時)",
    "出勤 (分)",
    "退勤 (時)",
    "退勤 (分)",
    "休暇種別",
)
HOUR_OPTIONS = tuple(str(hour) for hour in range(6, 23))
MINUTE_OPTIONS = ("00", "10", "20", "30", "40", "50")
LEAVE_OPTIONS = ("年休", "忌引")


def build_annual_workbook(fiscal_year: int) -> Workbook:
    workbook = Workbook()
    active_sheet = cast(Worksheet, workbook.active)
    workbook.remove(active_sheet)

    for month in FISCAL_MONTHS:
        worksheet = workbook.create_sheet(title=sheet_name(month))
        populate_month_sheet(worksheet, fiscal_year, month)

    return workbook


def build_monthly_workbook(fiscal_year: int, month: int) -> Workbook:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = sheet_name(month)
    populate_month_sheet(worksheet, fiscal_year, month)
    return workbook


def populate_month_sheet(worksheet: Worksheet, fiscal_year: int, month: int) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.append(list(HEADERS))

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for day in iter_month_dates(fiscal_year, month):
        worksheet.append([day, None, None, None, None, None])

    last_row = worksheet.max_row
    for row in range(2, last_row + 1):
        worksheet.cell(row=row, column=1).number_format = "d"

    _apply_column_widths(worksheet)
    _apply_validations(worksheet, last_row)


def save_workbook(workbook: Workbook, destination: Path) -> None:
    workbook.save(destination)


def ensure_output_paths_do_not_exist(paths: Iterable[Path]) -> None:
    for path in paths:
        if path.exists():
            raise FileExistsError(f"output path already exists: {path}")


def _apply_column_widths(worksheet: Worksheet) -> None:
    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 10
    worksheet.column_dimensions["D"].width = 10
    worksheet.column_dimensions["E"].width = 10
    worksheet.column_dimensions["F"].width = 12


def _apply_validations(worksheet: Worksheet, last_row: int) -> None:
    validations = (
        ("B", HOUR_OPTIONS),
        ("C", MINUTE_OPTIONS),
        ("D", HOUR_OPTIONS),
        ("E", MINUTE_OPTIONS),
        ("F", LEAVE_OPTIONS),
    )

    for column, options in validations:
        validation = DataValidation(
            type="list",
            formula1=_list_formula(options),
            allow_blank=True,
        )
        validation.add(f"{column}2:{column}{last_row}")
        worksheet.add_data_validation(validation)


def _list_formula(options: tuple[str, ...]) -> str:
    joined = ",".join(options)
    return f'"{joined}"'
